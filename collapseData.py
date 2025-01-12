# This function is utilized to use the collapse (sum) data from total_calories.py and paste it into an Excel sheet for the use of the professor

import pandas as pd
from openpyxl import load_workbook

def add_sheet_with_dataframe(sheet_name, dataframe):
    """
    Adds a new sheet with the given name and writes a Pandas DataFrame to it.
    
    Parameters:
    - file_path (str): Path to the Excel file.
    - sheet_name (str): Name of the new sheet.
    - dataframe (pd.DataFrame): DataFrame to write to the sheet.
    """
    try:

        file_path = 'data/calsPorSubbarrio.xlsx'

        # Load the existing workbook
        workbook = load_workbook(file_path)
        
        # Check if the sheet already exists
        if sheet_name in workbook.sheetnames:
            print(f"Sheet '{sheet_name}' already exists in the workbook.")
        else:
            # Use ExcelWriter with the existing workbook
            with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="new") as writer:
                dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
                print(f"Sheet '{sheet_name}' with the DataFrame has been added successfully.")
    except FileNotFoundError:
        print(f"The file '{file_path}' does not exist.")
    except Exception as e:
        print(f"An error occurred: {e}")